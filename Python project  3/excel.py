import openpyxl
import re

def validate_roll_number(roll_number):
    return bool(re.match(r'^[A-Za-z0-9]+$', roll_number))

def validate_name(name):
    return bool(re.match(r'^[A-Za-z0-9\s\-_]+$', name))

def validate_age(age):
    return age.isnumeric() and 0 <= int(age) <= 150

def validate_branch(branch):
    return bool(re.match(r'^[A-Za-z]+$', branch))

def validate_phone_number(phone_number):
    # We are assuming phone number should be numeric and have 10 digits.
    return phone_number.isnumeric() and len(phone_number) == 10

def add_data_to_excel(roll_number, name, age, branch, phone_number):
    # Load existing workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook('data.xlsx')
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the active sheet (first sheet by default)
    sheet = workbook.active

    # If the sheet is empty, add header row
    if sheet.max_row == 1:
        sheet.append(['Roll Number', 'Name', 'Age', 'Branch', 'Phone Number'])

    # Append the data to the sheet if all inputs are valid
    if (validate_roll_number(roll_number) and validate_name(name) and
            validate_age(age) and validate_branch(branch) and validate_phone_number(phone_number)):
        sheet.append([roll_number, name, age, branch, phone_number])
        # Save the workbook
        workbook.save('data.xlsx')
        print("Data saved successfully!")
    else:
        print("Invalid input. Roll number should contain only alphabets and numerics, "
              "name should contain alphabets, numerics, spaces, hyphens, and underscores, "
              "age should be a number between 0 and 150, "
              "branch should contain only alphabets, "
              "and phone number should be a 10-digit numeric value.")

def main():
    # Input roll number, name, age, branch, and phone number
    roll_number = input("Enter Roll Number: ")
    name = input("Enter Name: ")
    age = input("Enter Age: ")
    branch = input("Enter Branch: ")
    phone_number = input("Enter Phone Number: ")

    # Display successful message
    print("Successfully!")

    # Add data to Excel file if all inputs are valid
    add_data_to_excel(roll_number, name, age, branch, phone_number)

if __name__ == "__main__":
    main()
